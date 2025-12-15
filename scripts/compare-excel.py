#!/usr/bin/env python3
"""Compare two Excel files and report differences."""

import sys
import argparse

def compare_excel_files(reference_path: str, generated_path: str) -> int:
    """Compare two Excel files and return exit code (0=match, 1=mismatch)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("⚠️  openpyxl not installed. Run: pip install openpyxl")
        return 1

    try:
        ref_wb = load_workbook(reference_path, data_only=True)
        gen_wb = load_workbook(generated_path, data_only=True)
    except Exception as e:
        print(f"❌ Error loading workbooks: {e}")
        return 1

    print(f"  Reference sheets: {ref_wb.sheetnames}")
    print(f"  Generated sheets: {gen_wb.sheetnames}")
    print("")

    errors = []
    warnings = []

    # Check sheet count
    if len(ref_wb.sheetnames) != len(gen_wb.sheetnames):
        errors.append(
            f"Sheet count mismatch: reference={len(ref_wb.sheetnames)}, "
            f"generated={len(gen_wb.sheetnames)}"
        )

    # Compare each sheet
    for sheet_name in ref_wb.sheetnames:
        if sheet_name not in gen_wb.sheetnames:
            errors.append(f"Missing sheet: {sheet_name}")
            continue

        ref_sheet = ref_wb[sheet_name]
        gen_sheet = gen_wb[sheet_name]

        ref_rows = list(ref_sheet.iter_rows(values_only=True))
        gen_rows = list(gen_sheet.iter_rows(values_only=True))

        if len(ref_rows) != len(gen_rows):
            errors.append(
                f"Sheet {sheet_name}: row count mismatch - "
                f"reference={len(ref_rows)}, generated={len(gen_rows)}"
            )

        # Compare cell values
        min_rows = min(len(ref_rows), len(gen_rows))
        cell_diffs = 0

        for ri in range(min_rows):
            ref_row = ref_rows[ri] if ri < len(ref_rows) else ()
            gen_row = gen_rows[ri] if ri < len(gen_rows) else ()
            max_cols = max(
                len(ref_row) if ref_row else 0,
                len(gen_row) if gen_row else 0
            )

            for ci in range(max_cols):
                ref_val = ref_row[ci] if ref_row and ci < len(ref_row) else None
                gen_val = gen_row[ci] if gen_row and ci < len(gen_row) else None

                # Compare values (handle floating point tolerance)
                values_match = False
                if ref_val == gen_val:
                    values_match = True
                elif isinstance(ref_val, (int, float)) and isinstance(gen_val, (int, float)):
                    # Allow small floating point differences
                    if abs(ref_val - gen_val) < 0.0001:
                        values_match = True

                if not values_match:
                    cell_diffs += 1
                    if cell_diffs <= 5:
                        warnings.append(
                            f'  {sheet_name}[row {ri+1}, col {ci+1}]: '
                            f'"{ref_val}" vs "{gen_val}"'
                        )

        if cell_diffs > 0:
            errors.append(f"Sheet {sheet_name}: {cell_diffs} cell value differences")
            if cell_diffs > 5:
                warnings.append(f"  ... and {cell_diffs - 5} more differences")

    # Check for extra sheets in generated file
    for sheet_name in gen_wb.sheetnames:
        if sheet_name not in ref_wb.sheetnames:
            errors.append(f"Extra sheet in generated: {sheet_name}")

    # Print results
    if errors:
        print("❌ Comparison FAILED:")
        for e in errors:
            print(f"  - {e}")
        if warnings:
            print("")
            print("First differences:")
            for w in warnings:
                print(w)
        return 1
    else:
        print("✅ Comparison PASSED: Generated report matches reference")
        return 0


def main():
    parser = argparse.ArgumentParser(
        description="Compare two Excel files and report differences"
    )
    parser.add_argument(
        "reference",
        help="Path to reference Excel file"
    )
    parser.add_argument(
        "generated",
        help="Path to generated Excel file"
    )
    args = parser.parse_args()

    sys.exit(compare_excel_files(args.reference, args.generated))


if __name__ == "__main__":
    main()
